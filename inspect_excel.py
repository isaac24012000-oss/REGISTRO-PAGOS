#!/usr/bin/env python3
"""
Script para inspeccionar el archivo Excel
"""

try:
    import openpyxl
    
    archivo = r'C:\Users\USUARIO\Desktop\REGISTRO DE PAGOS\DATA ENERO 2026.xlsx'
    print(f"Leyendo archivo: {archivo}\n")
    
    wb = openpyxl.load_workbook(archivo)
    ws = wb.active
    
    print("Columnas:")
    encabezados = [cell.value for cell in ws[1]]
    for i, col in enumerate(encabezados, 1):
        print(f"  {i}. {col}")
    
    print(f"\nTotal de filas: {ws.max_row}")
    print(f"\nPrimeras 5 filas de datos:")
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), 1):
        print(f"Fila {row_idx}: {row}")
    
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
