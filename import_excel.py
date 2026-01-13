#!/usr/bin/env python3
"""
Script para importar datos del Excel a la base de datos
"""

import openpyxl
from database import init_db, registrar_pago_planilla, registrar_gasto_administrativo
from datetime import datetime
import os

def importar_excel_a_bd(archivo_excel=None):
    """Importa datos del Excel a la base de datos"""
    
    # Si no se especifica ruta, usar ruta relativa
    if archivo_excel is None:
        archivo_excel = os.path.join(os.path.dirname(__file__), 'DATA ENERO 2026.xlsx')
    
    print("=" * 70)
    print("IMPORTADOR DE DATOS EXCEL - SISTEMA DE REGISTRO DE PAGOS")
    print("=" * 70)
    print()
    
    # Inicializar BD
    init_db()
    
    try:
        print(f"Leyendo archivo: {archivo_excel}\n")
        
        wb = openpyxl.load_workbook(archivo_excel)
        ws = wb.active
        
        rucs_procesados = set()
        pagos_count = 0
        
        # Procesar filas (saltando encabezado)
        total_filas = ws.max_row - 1
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
            try:
                campana = row[0]
                documento = str(int(row[1])) if row[1] else None
                razon_social = row[2]
                deuda_total = float(row[3]) if row[3] else 0
                gastos_admin = float(row[4]) if row[4] else 0
                periodos = row[5]
                asesor = row[6]
                
                if not documento or not razon_social or not campana:
                    continue
                
                # Usar RUC como ID (ya que todos son unicos)
                ruc_id = documento
                
                # Registrar UN UNICO pago: elegir el monto mas alto
                if deuda_total > 0 or gastos_admin > 0:
                    if deuda_total >= gastos_admin:
                        # Registrar como PLANILLA (tiene mas deuda)
                        registrar_pago_planilla(
                            ruc_id, documento, campana, deuda_total, 
                            "PENDIENTE",
                            fecha_promesa=None,
                            observaciones=f"Periodos: {periodos} | Asesor: {asesor}"
                        )
                    else:
                        # Registrar como GASTO_ADMINISTRATIVO (tiene mas gasto)
                        registrar_gasto_administrativo(
                            ruc_id, documento, campana, gastos_admin,
                            "PENDIENTE",
                            fecha_promesa=None,
                            observaciones=f"Periodos: {periodos} | Asesor: {asesor}"
                        )
                    pagos_count += 1
                    rucs_procesados.add(documento)
                
                if row_idx % 500 == 0:
                    print(f"  Procesadas {row_idx}/{total_filas} filas...")
            
            except Exception as e:
                print(f"  Error en fila {row_idx}: {str(e)}")
                continue
        
        print()
        print("=" * 70)
        print("IMPORTACION COMPLETADA")
        print("=" * 70)
        print(f"RUCs unicos procesados: {len(rucs_procesados)}")
        print(f"Pagos unicos registrados: {pagos_count}")
        print()
        
        return True
    
    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    importar_excel_a_bd()
