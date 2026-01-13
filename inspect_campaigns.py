import openpyxl

wb = openpyxl.load_workbook('DATA ENERO 2026.xlsx', data_only=True)
ws = wb.active

# Get headers
headers = [cell.value for cell in ws[1]]
print("📋 Columnas del Excel:")
for i, h in enumerate(headers, 1):
    print(f"  {i:2d}. {h}")

# Get unique values in possible campaign columns
print("\n🔍 Buscando campaña...")

# Check for campaign-related columns
campaign_indices = []
for i, h in enumerate(headers):
    if h and ('campaña' in str(h).lower() or 'campaign' in str(h).lower() or 'flujo' in str(h).lower()):
        campaign_indices.append(i)
        print(f"  Columna {i+1} ({h}) podría ser campaña")

# Get first 10 rows
print("\n📊 Primeros 10 registros:")
for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=11, values_only=True), 2):
    print(f"  Row {row_num}: {row[:8]}")

# Get unique values in each column to find campaigns
print("\n🔢 Valores únicos por columna (primeras 30 filas):")
for col_idx in range(len(headers)):
    values = set()
    for row in ws.iter_rows(min_row=2, max_row=31, values_only=True):
        val = row[col_idx]
        if val:
            values.add(str(val).strip())
    if len(values) <= 10:  # Only show columns with few unique values
        print(f"  Columna {col_idx+1} ({headers[col_idx]}): {sorted(values)}")
